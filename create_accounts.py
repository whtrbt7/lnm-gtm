"""
GTM Account + Container bulk creator.

For each Tier 2 row in the XLSX that has no GTM ID (Col AC):
  1. Uses Playwright to create a GTM Account + Web container via the GTM UI.
  2. Captures the new GTM-XXXXXXXX ID from the URL after creation.
  3. Writes the ID back to Col AC in the XLSX.

Account name  : "{Client Name} - {Folder Name}"  (Col B - Col C)
Container name: Client URL                         (Col U)
Container type: Web
"""

import re
import time
import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from auth import get_gtm_service
from utils import clean_url

import os as _os
XLSX_PATH = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'GTM Bulk Setup OktoRocket.xlsx')
SHEET_NAME = 'AA Client Import List (1)'
GTM_URL = 'https://tagmanager.google.com/'

# Column indices (0-based)
COL_TIER     = 0   # A
COL_NAME     = 1   # B
COL_FOLDER   = 2   # C
COL_URL      = 20  # U
COL_GTM_DONE = 19  # T — Alex GAds & GTM Setup date
COL_GTM_ID   = 28  # AC


def load_rows_needing_accounts(ws, limit=None):
    """Return rows with no GTM ID and no GTM done date — regardless of tier."""
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[COL_NAME]
        if not name:
            continue
        gtm_done = str(row[COL_GTM_DONE] or '').strip()
        if gtm_done and gtm_done not in ('None', 'N/A', 'main site skip'):
            continue  # already done
        raw_id = row[COL_GTM_ID]
        gtm_id = raw_id if isinstance(raw_id, str) and raw_id.startswith('GTM-') else None
        folder = row[COL_FOLDER]
        url    = row[COL_URL]
        rows.append((i, name, folder, url, gtm_id))
        if limit and len(rows) >= limit:
            break
    return rows


def load_rows_for_tiers(ws, tiers, limit=None):
    """Return list of (row_number, name, folder, url, gtm_id) for the given tier values."""
    # Normalise tiers to strings for comparison (handles 3.0, '3', 'Not Found', etc.)
    tier_set = {str(t).strip() for t in tiers}
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tier = row[COL_TIER]
        if str(tier).strip() not in tier_set:
            continue
        name    = row[COL_NAME]
        folder  = row[COL_FOLDER]
        url     = row[COL_URL]
        raw_id  = row[COL_GTM_ID]
        gtm_id  = raw_id if isinstance(raw_id, str) and raw_id.startswith('GTM-') else None
        rows.append((i, name, folder, url, gtm_id))
        if limit and len(rows) >= limit:
            break
    return rows


def create_gtm_account(context, account_name, container_name):
    """
    Opens a fresh page, creates a GTM account + Web container, then closes the page.
    Returns the GTM public ID (e.g. 'GTM-XXXXXX') or None on failure.
    """
    page = context.new_page()
    try:
        print(f'  -> Navigating to create account page...')
        page.goto(f'{GTM_URL}#/home', wait_until='domcontentloaded', timeout=60000)
        time.sleep(2)

        # Click "Create Account" button
        page.get_by_role('button', name=re.compile('Create Account', re.I)).first.click(timeout=10000)
        page.wait_for_url(re.compile(r'admin/accounts/create'), timeout=10000)
        time.sleep(1)

        # --- Account name ---
        account_field = page.locator('input[name="form.account.properties.displayName"]')
        account_field.wait_for(timeout=10000)
        account_field.fill(account_name)

        # --- Container name ---
        container_field = page.locator('input[name="form.container.properties.displayName"]')
        container_field.fill(container_name)

        # --- Select Web platform ---
        page.locator('div.context-picker__row[title="Web"]').click(timeout=10000)
        time.sleep(1)

        # --- Click Create ---
        page.get_by_role('button', name=re.compile(r'^Create$')).click(timeout=10000)

        # Snapshot all workspace URLs already open BEFORE TOS/redirect
        pre_workspace_urls = {
            p.url for p in context.pages
            if re.search(r'tagmanager\.google\.com.*workspaces', p.url)
        }

        # --- Accept TOS if it appears ---
        try:
            tos_agree = page.get_by_role('button', name=re.compile('Yes|I agree|Accept', re.I))
            tos_agree.wait_for(timeout=8000)
            tos_agree.click()
        except PlaywrightTimeoutError:
            pass  # No TOS dialog on this page

        # After TOS, find the NEW workspace page (not in pre_workspace_urls)
        workspace_page = None
        for _ in range(25):
            time.sleep(1)
            for p in context.pages:
                url = p.url
                if re.search(r'tagmanager\.google\.com.*workspaces', url) and url not in pre_workspace_urls:
                    workspace_page = p
                    break
            if workspace_page:
                break

        if not workspace_page:
            # Fallback: current page may have navigated to workspace
            try:
                page.wait_for_url(re.compile(r'tagmanager\.google\.com.*workspaces'), timeout=10000)
                if page.url not in pre_workspace_urls:
                    workspace_page = page
            except PlaywrightTimeoutError:
                pass

        if not workspace_page:
            print(f'  WARNING: Could not find new workspace page.')
            return None

        time.sleep(2)
        workspace_url = workspace_page.url
        url_match = re.search(r'accounts/(\d+)/containers/(\d+)', workspace_url)

        # Try to extract GTM public ID directly from the rendered page before closing
        gtm_public_id = None
        try:
            workspace_page.wait_for_load_state('networkidle', timeout=15000)
            # Use innerText (visible rendered text) to find GTM-XXXXXX — reliable on Angular SPA
            gtm_public_id = workspace_page.evaluate(
                "() => { const m = document.body.innerText.match(/GTM-[A-Z0-9]{4,10}/); return m ? m[0] : null; }"
            )
            if gtm_public_id:
                print(f'  Got GTM ID from page: {gtm_public_id}')
            else:
                # Fallback: search raw HTML (catches IDs in attributes/scripts)
                id_match = re.search(r'GTM-[A-Z0-9]{4,10}', workspace_page.content())
                if id_match:
                    gtm_public_id = id_match.group(0)
                    print(f'  Got GTM ID from HTML: {gtm_public_id}')
        except Exception as e:
            print(f'  Could not extract GTM ID from page: {e}')

        # Close the workspace page to prevent stale URL accumulation
        try:
            workspace_page.close()
        except Exception:
            pass

        if gtm_public_id:
            return gtm_public_id

        # Fallback: API lookup by account/container numeric IDs
        if url_match:
            account_id   = url_match.group(1)
            container_id = url_match.group(2)
            try:
                service = get_gtm_service()
                container = service.accounts().containers().get(
                    path=f'accounts/{account_id}/containers/{container_id}'
                ).execute()
                return container.get('publicId')
            except Exception as e:
                print(f'  API lookup failed: {e}')

        print(f'  Workspace URL: {workspace_url}')
        return None

    finally:
        try:
            page.close()
        except Exception:
            pass


def run(dry_run=False, limit=None, tiers=None):
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    if tiers:
        rows = load_rows_for_tiers(ws, tiers, limit=limit)
        tier_label = ', '.join(str(t) for t in tiers)
    else:
        rows = load_rows_needing_accounts(ws, limit=limit)
        tier_label = 'all (no GTM ID + not done)'

    to_create = [(rn, n, f, clean_url(u), g) for rn, n, f, u, g in rows if not g]
    skipped   = [(rn, n, f, u, g) for rn, n, f, u, g in rows if g]

    print(f'Tiers: {tier_label} — {len(rows)} rows found:')
    print(f'  {len(skipped)} already have GTM IDs — will skip')
    print(f'  {len(to_create)} need creation\n')

    for rn, name, folder, url, _ in skipped:
        print(f'  SKIP  Row {rn}: {name} - {folder}')

    if dry_run:
        print('\n[DRY RUN] Would create:')
        for rn, name, folder, url, _ in to_create:
            account_name = f'{name} - {folder}'
            print(f'  Row {rn}: Account="{account_name}" | Container="{url}" | Type=WEB')
        return

    print()
    # Connect to Chrome launched with --remote-debugging-port=9222
    # Launch command:
    #   /Applications/Google\ Chrome.app/Contents/MacOS/Google\ Chrome \
    #     --remote-debugging-port=9222 --user-data-dir=/tmp/chrome_gtm_auto \
    #     --no-first-run https://tagmanager.google.com/
    CDP_URL = 'http://localhost:9222'

    def get_context(p):
        """Connect to Chrome via CDP, waiting if it's not ready yet."""
        for attempt in range(10):
            try:
                browser = p.chromium.connect_over_cdp(CDP_URL)
                context = browser.contexts[0]
                return browser, context
            except Exception as e:
                print(f'  CDP connect failed (attempt {attempt+1}): {e}')
                print(f'  Make sure Chrome is running with --remote-debugging-port=9222. Retrying in 10s...')
                time.sleep(10)
        raise RuntimeError('Could not connect to Chrome after 10 attempts.')

    with sync_playwright() as p:
        browser, context = get_context(p)

        # Verify we're logged in
        print('Connecting to Chrome via CDP...')
        check_page = context.new_page()
        check_page.goto(GTM_URL, wait_until='domcontentloaded', timeout=60000)
        if 'accounts.google.com' in check_page.url:
            print('Please log in to Google in the Chrome window (use Authenticator app for 2FA).')
            print('Waiting up to 3 minutes...')
            check_page.wait_for_url(re.compile(r'tagmanager\.google\.com(?!/.*signin)'), timeout=180000)
        check_page.close()

        print('Logged in. Starting account creation...')
        time.sleep(3)

        for rn, name, folder, url, _ in to_create:
            account_name   = f'{name} - {folder}'
            container_name = url or name
            print(f'\nCreating: "{account_name}" / "{container_name}"')

            # Reconnect if browser/context was closed
            try:
                _ = context.pages
            except Exception:
                print('  Browser closed, reconnecting...')
                browser, context = get_context(p)

            try:
                gtm_id = create_gtm_account(context, account_name, container_name)
                if gtm_id:
                    print(f'  SUCCESS: {gtm_id}')
                    # Save via a fresh workbook load to preserve formulas
                    wb_write = openpyxl.load_workbook(XLSX_PATH)
                    wb_write[SHEET_NAME].cell(row=rn, column=COL_GTM_ID + 1).value = gtm_id
                    wb_write.save(XLSX_PATH)
                    print(f'  Saved {gtm_id} to row {rn} in XLSX.')
                else:
                    print(f'  WARNING: Could not capture GTM ID for row {rn}.')
            except Exception as e:
                print(f'  ERROR on row {rn}: {e}')

            time.sleep(2)

        browser.close()

    print('\nDone.')


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true', help='Preview without creating')
    parser.add_argument('--limit', type=int, default=None, help='Max rows to process')
    parser.add_argument('--tiers', nargs='+', default=None, help='Filter by tier (e.g. --tiers 2 3). Omit to process all rows with no GTM ID and no done date.')
    args = parser.parse_args()

    run(dry_run=args.dry_run, limit=args.limit, tiers=args.tiers)
