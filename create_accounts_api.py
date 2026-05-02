"""
GTM Account + Container bulk creator — API version.

Replaces the Playwright-based create_accounts.py with direct GTM API v2 calls.
For each row with no GTM ID and no done date, creates a GTM Account + Web container
and writes the GTM public ID back to col 28 (AC) in the XLSX.

Account name  : "{Client Name} - {Folder Name}"
Container name: Client URL (col U)
Container type: Web

Usage:
    python create_accounts_api.py               # All rows needing creation
    python create_accounts_api.py --dry-run     # Preview only
    python create_accounts_api.py --limit 10    # Process first N rows
    python create_accounts_api.py --token-file token_analytics.json  # default
"""

import json
import os
import time
import argparse
import openpyxl
from googleapiclient.errors import HttpError
from utils import clean_url

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH   = os.path.join(SCRIPT_DIR, 'GTM Bulk Setup OktoRocket.xlsx')
SHEET_NAME  = 'AA Client Import List (1)'

COL_NAME     = 1
COL_FOLDER   = 2
COL_URL      = 20
COL_GTM_DONE = 19
COL_GTM_ID   = 28


def get_gtm_service(token_file='token_analytics.json'):
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    path = os.path.join(SCRIPT_DIR, token_file)
    with open(path) as f:
        data = json.load(f)

    creds = Credentials(
        token=data.get('token'),
        refresh_token=data.get('refresh_token'),
        token_uri=data.get('token_uri', 'https://oauth2.googleapis.com/token'),
        client_id=data.get('client_id'),
        client_secret=data.get('client_secret'),
        scopes=data.get('scopes'),
    )
    if not creds.valid:
        if creds.expired and creds.refresh_token:
            print('  [auth] Refreshing token...')
            creds.refresh(Request())
            data['token'] = creds.token
            data['expiry'] = creds.expiry.isoformat() if creds.expiry else None
            with open(path, 'w') as f:
                json.dump(data, f, indent=2)
        else:
            raise RuntimeError('Token invalid and cannot be refreshed.')
    return build('tagmanager', 'v2', credentials=creds)


def _api_call_with_retry(call, max_retries=8, base_delay=3.0):
    delay = base_delay
    for attempt in range(max_retries):
        try:
            return call()
        except HttpError as e:
            status = e.resp.status
            if status in (429, 500, 503) and attempt < max_retries - 1:
                print(f'  [retry] HTTP {status}, waiting {delay:.1f}s...')
                time.sleep(delay)
                delay = min(delay * 2, 60)
            else:
                raise


def load_rows_needing_accounts(ws, limit=None):
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[COL_NAME]
        if not name:
            continue
        gtm_done = str(row[COL_GTM_DONE] or '').strip()
        if gtm_done and gtm_done not in ('None', 'N/A', 'main site skip'):
            continue
        raw_id = row[COL_GTM_ID]
        gtm_id = raw_id if isinstance(raw_id, str) and raw_id.startswith('GTM-') else None
        if gtm_id:
            continue  # already has a GTM ID, skip
        folder = row[COL_FOLDER]
        url    = row[COL_URL]
        rows.append((i, name, folder, url))
        if limit and len(rows) >= limit:
            break
    return rows


def writeback_gtm_id(row_num, gtm_id):
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]
    ws.cell(row=row_num, column=COL_GTM_ID + 1).value = gtm_id
    wb.save(XLSX_PATH)


def run(dry_run=False, limit=None, token_file='token_analytics.json'):
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    rows = load_rows_needing_accounts(ws, limit=limit)
    print(f'{len(rows)} rows need GTM account creation\n')

    if dry_run:
        for rn, name, folder, url in rows:
            account_name   = f'{name} - {folder}' if folder else name
            container_name = clean_url(url) or name
            print(f'  Row {rn}: Account="{account_name}" | Container="{container_name}"')
        return

    service = get_gtm_service(token_file)

    success = 0
    failed  = 0

    for rn, name, folder, url in rows:
        account_name   = f'{name} - {folder}' if folder else name
        container_name = clean_url(url) or name

        print(f'Row {rn}: "{account_name}" / "{container_name}"')
        try:
            # 1. Create GTM account
            account = _api_call_with_retry(lambda an=account_name: service.accounts().create(
                body={'name': an}
            ).execute())
            account_id = account['accountId']

            # 2. Create Web container inside it
            container = _api_call_with_retry(lambda aid=account_id, cn=container_name:
                service.accounts().containers().create(
                    parent=f'accounts/{aid}',
                    body={'name': cn, 'usageContext': ['WEB']}
                ).execute())

            gtm_id = container.get('publicId')
            print(f'  SUCCESS: {gtm_id} (account {account_id})')

            writeback_gtm_id(rn, gtm_id)
            success += 1

            # Throttle to stay under quota
            time.sleep(1.5)

        except Exception as e:
            print(f'  ERROR: {e}')
            failed += 1

    print(f'\n=== Done: {success} created, {failed} failed ===')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run',    action='store_true')
    parser.add_argument('--limit',      type=int, default=None)
    parser.add_argument('--token-file', default='token_analytics.json')
    args = parser.parse_args()

    run(dry_run=args.dry_run, limit=args.limit, token_file=args.token_file)
