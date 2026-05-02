"""
GTM Bulk Push — LNM Standard v1.0

Reads the XLSX and pushes tags/triggers directly into each client's live GTM
container via the GTM API v2.  Containers are matched by URL (col 20) or
GTM public ID (col 28).

Usage:
    python push_gtm_setup.py                  # All rows with data
    python push_gtm_setup.py --tier 2         # Tier 2 only
    python push_gtm_setup.py --row 138        # Single row
    python push_gtm_setup.py --dry-run        # Preview matches + planned actions
    python push_gtm_setup.py --force-recreate # Delete and recreate existing items
"""

import re
import os
import sys
import json
import time
import argparse
import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
TOKEN_PATH    = os.path.join(SCRIPT_DIR, 'token.json')
INDEX_CACHE   = os.path.join(SCRIPT_DIR, 'container_index_cache.json')
XLSX_PATH     = os.path.join(SCRIPT_DIR, 'GTM Bulk Setup OktoRocket.xlsx')
SHEET_NAME    = 'AA Client Import List (1)'

# ── Column indices (0-based) ───────────────────────────────────────────────────
COL_TIER        = 0
COL_NAME        = 1
COL_URL         = 20   # client website URL (col 21 in sheet, 0-based index 20)
COL_SCHED       = 21
COL_PHONE       = 22
COL_GA4_ID      = 23
COL_GADS_ID     = 25
COL_LABEL_APPT  = 26
COL_LABEL_PHONE = 27
COL_GTM_PUBLIC  = 28   # GTM-XXXXXXX public ID, col 28 (0-based)
COL_MULTI_PHONE = 32   # JSON array of extra phone numbers
COL_GTM_DONE    = 19   # Alex GAds & GTM Setup — date written here when complete


# ── Auth ──────────────────────────────────────────────────────────────────────

def get_gtm_service(token_path=None):
    """Load credentials, refresh if needed, return GTM API v2 service."""
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    path = token_path or TOKEN_PATH
    with open(path) as f:
        data = json.load(f)

    creds = Credentials(
        token=data.get('token'),
        refresh_token=data.get('refresh_token'),
        token_uri=data.get('token_uri', 'https://oauth2.googleapis.com/token'),
        client_id=data.get('client_id'),
        client_secret=data.get('client_secret'),
        scopes=data.get('scopes'),
    )

    if not creds.valid:
        if creds.expired and creds.refresh_token:
            print("  [auth] Refreshing expired token...")
            creds.refresh(Request())
            # Persist refreshed token
            data['token'] = creds.token
            data['expiry'] = creds.expiry.isoformat() if creds.expiry else None
            with open(path, 'w') as f:
                json.dump(data, f, indent=2)
        else:
            raise RuntimeError("Credentials invalid and cannot be refreshed. Re-authenticate.")

    return build('tagmanager', 'v2', credentials=creds)


# ── Container index ───────────────────────────────────────────────────────────

def _normalize_domain(url):
    """Strip protocol/www/trailing-slash from URL to get bare domain."""
    if not url:
        return ''
    s = str(url).strip().lower()
    s = re.sub(r'^https?://', '', s)
    s = re.sub(r'^www\.', '', s)
    s = s.rstrip('/')
    # strip path, keep domain only
    s = s.split('/')[0]
    return s


def _api_call_with_retry(call, max_retries=8, base_delay=3.0):
    """Execute a callable that makes one API request, retrying on 429/5xx."""
    from googleapiclient.errors import HttpError
    delay = base_delay
    for attempt in range(max_retries):
        try:
            return call()
        except HttpError as e:
            status = e.resp.status
            if status in (429, 500, 503) and attempt < max_retries - 1:
                print(f"  [retry] HTTP {status}, waiting {delay:.1f}s...")
                time.sleep(delay)
                delay = min(delay * 2, 60)
            else:
                raise


def load_cached_index(cache_path=None):
    """Load url_index/gtm_index from disk cache if present."""
    path = cache_path or INDEX_CACHE
    if not os.path.exists(path):
        return None, None
    with open(path) as f:
        data = json.load(f)
    url_index = {k: tuple(v) for k, v in data.get('url_index', {}).items()}
    gtm_index = {k: tuple(v) for k, v in data.get('gtm_index', {}).items()}
    print(f"  [cache] Loaded {len(url_index)} url / {len(gtm_index)} GTM entries "
          f"from {path}")
    return url_index, gtm_index


def save_cached_index(url_index, gtm_index, cache_path=None):
    path = cache_path or INDEX_CACHE
    with open(path, 'w') as f:
        json.dump({'url_index': url_index, 'gtm_index': gtm_index}, f)
    print(f"  [cache] Saved index to {path}")


def build_container_index(service, cache_path=None):
    """
    Return two dicts:
      url_index : {normalized_domain → (account_id, container_id, public_id)}
      gtm_index : {GTM-XXXXXXX → (account_id, container_id, public_id)}
    """
    url_index = {}
    gtm_index = {}

    accounts_resp = _api_call_with_retry(lambda: service.accounts().list().execute())
    accounts = accounts_resp.get('account', [])
    print(f"  [index] Found {len(accounts)} GTM account(s)")

    for idx, acct in enumerate(accounts, 1):
        acct_path = acct['path']           # e.g. "accounts/123456"
        acct_id   = acct['accountId']

        if idx % 20 == 0:
            print(f"  [index] Scanning account {idx}/{len(accounts)}...")

        ctrs_resp = _api_call_with_retry(
            lambda p=acct_path: service.accounts().containers().list(parent=p).execute()
        )
        containers = ctrs_resp.get('container', [])

        for ctr in containers:
            ctr_id    = ctr['containerId']
            public_id = ctr.get('publicId', '')   # GTM-XXXXXXX
            name      = ctr.get('name', '')        # usually the domain

            domain = _normalize_domain(name)
            entry  = (acct_id, ctr_id, public_id)

            if domain:
                url_index[domain] = entry
            if public_id:
                gtm_index[public_id] = entry

        # Throttle: ~1 request/sec to stay well under the per-minute quota
        time.sleep(1.2)

    print(f"  [index] Indexed {len(url_index)} containers by domain, "
          f"{len(gtm_index)} by GTM ID")
    save_cached_index(url_index, gtm_index, cache_path)
    return url_index, gtm_index


# ── Container matching ────────────────────────────────────────────────────────

def find_container(url_index, gtm_index, row_url, row_gtm_id):
    """
    Return (account_id, container_id, public_id) or None.
    Prefers GTM ID lookup; falls back to domain.
    """
    gtm_raw = str(row_gtm_id or '').strip()
    if gtm_raw and gtm_raw.upper().startswith('GTM-'):
        entry = gtm_index.get(gtm_raw.upper())
        if entry:
            return entry

    domain = _normalize_domain(row_url)
    if domain:
        return url_index.get(domain)

    return None


# ── Workspace ─────────────────────────────────────────────────────────────────

def get_workspace(service, account_id, container_id):
    """Return the workspace_id of the first (Default) workspace."""
    parent = f"accounts/{account_id}/containers/{container_id}"
    resp = _api_call_with_retry(
        lambda: service.accounts().containers().workspaces().list(parent=parent).execute()
    )
    workspaces = resp.get('workspace', [])
    if not workspaces:
        raise RuntimeError(f"No workspaces found for container {parent}")
    ws = workspaces[0]
    return ws['workspaceId']


# ── Idempotency ───────────────────────────────────────────────────────────────

def get_existing_triggers(service, acct, ctr, ws):
    """Return {trigger_name: trigger_id}."""
    parent = f"accounts/{acct}/containers/{ctr}/workspaces/{ws}"
    resp = _api_call_with_retry(
        lambda: service.accounts().containers().workspaces().triggers().list(
            parent=parent).execute()
    )
    return {t['name']: t['triggerId'] for t in resp.get('trigger', [])}


def get_existing_tags(service, acct, ctr, ws):
    """Return {tag_name: tag_id}."""
    parent = f"accounts/{acct}/containers/{ctr}/workspaces/{ws}"
    resp = _api_call_with_retry(
        lambda: service.accounts().containers().workspaces().tags().list(
            parent=parent).execute()
    )
    return {t['name']: t['tagId'] for t in resp.get('tag', [])}


# ── Trigger / Tag creation ────────────────────────────────────────────────────

def ensure_trigger(service, acct, ctr, ws, trigger_body, existing_triggers,
                   force_recreate=False):
    """
    Create trigger if it doesn't exist (or force_recreate).
    Returns (trigger_id, status) where status is 'new'|'existed'|'recreated'.
    """
    name   = trigger_body['name']
    parent = f"accounts/{acct}/containers/{ctr}/workspaces/{ws}"

    if name in existing_triggers:
        if not force_recreate:
            return existing_triggers[name], 'existed'
        # Delete first
        tid = existing_triggers[name]
        _api_call_with_retry(
            lambda p=f"{parent}/triggers/{tid}":
                service.accounts().containers().workspaces().triggers().delete(path=p).execute()
        )

    # Strip accountId/containerId from body — API fills them in
    body = {k: v for k, v in trigger_body.items()
            if k not in ('accountId', 'containerId', 'triggerId')}

    result = _api_call_with_retry(
        lambda: service.accounts().containers().workspaces().triggers().create(
            parent=parent, body=body).execute()
    )
    status = 'recreated' if name in existing_triggers else 'new'
    return result['triggerId'], status


def ensure_tag(service, acct, ctr, ws, tag_body, existing_tags,
               force_recreate=False):
    """
    Create tag if it doesn't exist (or force_recreate).
    Returns (tag_id, status) where status is 'new'|'existed'|'recreated'.
    """
    name   = tag_body['name']
    parent = f"accounts/{acct}/containers/{ctr}/workspaces/{ws}"

    if name in existing_tags:
        if not force_recreate:
            return existing_tags[name], 'existed'
        tid = existing_tags[name]
        _api_call_with_retry(
            lambda p=f"{parent}/tags/{tid}":
                service.accounts().containers().workspaces().tags().delete(path=p).execute()
        )

    body = {k: v for k, v in tag_body.items()
            if k not in ('accountId', 'containerId', 'tagId')}

    result = _api_call_with_retry(
        lambda: service.accounts().containers().workspaces().tags().create(
            parent=parent, body=body).execute()
    )
    status = 'recreated' if name in existing_tags else 'new'
    return result['tagId'], status


# ── Trigger / Tag body builders ───────────────────────────────────────────────
# (Adapted from generate_gtm_exports.py — same structure, no dummy IDs)

def build_appt_trigger_body(sched_label, appt_event):
    return {
        "name": f"CE - {sched_label} - Appointment Booked",
        "type": "CUSTOM_EVENT",
        "customEventFilter": [{
            "type": "EQUALS",
            "parameter": [
                {"type": "TEMPLATE", "key": "arg0", "value": "{{_event}}"},
                {"type": "TEMPLATE", "key": "arg1", "value": appt_event}
            ]
        }]
    }


def build_cl_trigger_body(phone):
    return {
        "name": f"CL - Phone Click - {phone}",
        "type": "LINK_CLICK",
        "filter": [{
            "type": "CONTAINS",
            "parameter": [
                {"type": "TEMPLATE", "key": "arg0", "value": "{{Click URL}}"},
                {"type": "TEMPLATE", "key": "arg1", "value": phone}
            ]
        }],
        "parameter": [
            {"type": "BOOLEAN",  "key": "waitForTags",        "value": "true"},
            {"type": "BOOLEAN",  "key": "checkValidation",     "value": "true"},
            {"type": "TEMPLATE", "key": "waitForTagsTimeout",  "value": "2000"}
        ]
    }


def build_all_pages_trigger_body():
    return {"name": "All Pages", "type": "PAGEVIEW"}


def build_ga4_config_tag_body(ga4_id, all_pages_trigger_id):
    return {
        "name": "GA4 - Configuration",
        "type": "gaawc",
        "parameter": [
            {"type": "TEMPLATE", "key": "measurementId", "value": ga4_id}
        ],
        "firingTriggerId": [all_pages_trigger_id],
        "tagFiringOption": "ONCE_PER_EVENT",
        "monitoringMetadata": {"type": "MAP"},
        "consentSettings": {"consentStatus": "NOT_SET"}
    }


def build_ga4_event_appt_tag_body(ga4_id, appt_event, appt_trigger_id):
    return {
        "name": f"GA4 - Event - {appt_event}",
        "type": "gaawe",
        "parameter": [
            {"type": "TAG_REFERENCE", "key": "gaSettings", "value": "GA4 - Configuration"},
            {"type": "TEMPLATE",      "key": "eventName",  "value": appt_event},
        ],
        "firingTriggerId": [appt_trigger_id],
        "tagFiringOption": "ONCE_PER_EVENT",
        "monitoringMetadata": {"type": "MAP"},
        "consentSettings": {"consentStatus": "NOT_SET"}
    }


def build_ga4_event_phone_tag_body(ga4_id, cl_trigger_ids):
    return {
        "name": "GA4 - Event - phone_click",
        "type": "gaawe",
        "parameter": [
            {"type": "TAG_REFERENCE", "key": "gaSettings", "value": "GA4 - Configuration"},
            {"type": "TEMPLATE",      "key": "eventName",  "value": "phone_click"},
        ],
        "firingTriggerId": cl_trigger_ids,
        "tagFiringOption": "ONCE_PER_EVENT",
        "monitoringMetadata": {"type": "MAP"},
        "consentSettings": {"consentStatus": "NOT_SET"}
    }


def build_conversion_linker_tag_body(all_pages_trigger_id):
    return {
        "name": "Conversion Linker",
        "type": "gclidw",
        "parameter": [
            {"type": "BOOLEAN", "key": "enableCrossDomainLinking", "value": "false"},
            {"type": "BOOLEAN", "key": "enableUrlPassthrough",     "value": "false"},
            {"type": "BOOLEAN", "key": "decorateFormsWithData",    "value": "false"},
        ],
        "firingTriggerId": [all_pages_trigger_id],
        "tagFiringOption": "ONCE_PER_LOAD",
        "monitoringMetadata": {"type": "MAP"},
        "consentSettings": {"consentStatus": "NOT_SET"}
    }


def build_gads_appt_tag_body(store_name, gads_id_str, appt_label, appt_trigger_id):
    return {
        "name": f"GAds - {store_name} - Booked_Appointment",
        "type": "awct",
        "parameter": [
            {"type": "INTEGER",  "key": "conversionId",    "value": gads_id_str},
            {"type": "TEMPLATE", "key": "conversionLabel", "value": appt_label},
            {"type": "TEMPLATE", "key": "conversionValue", "value": "65"},
            {"type": "TEMPLATE", "key": "currencyCode",    "value": "USD"},
            {"type": "BOOLEAN",  "key": "remarketingOnly", "value": "false"},
            {"type": "BOOLEAN",  "key": "enabledMd",       "value": "true"}
        ],
        "firingTriggerId": [appt_trigger_id],
        "tagFiringOption": "ONCE_PER_EVENT",
        "monitoringMetadata": {"type": "MAP"},
        "consentSettings": {"consentStatus": "NOT_SET"}
    }


def build_gads_phone_tag_body(store_name, gads_id_str, phone_label, cl_trigger_ids):
    return {
        "name": f"GAds - {store_name} - Phone_Click",
        "type": "awct",
        "parameter": [
            {"type": "INTEGER",  "key": "conversionId",    "value": gads_id_str},
            {"type": "TEMPLATE", "key": "conversionLabel", "value": phone_label},
            {"type": "TEMPLATE", "key": "conversionValue", "value": "10"},
            {"type": "TEMPLATE", "key": "currencyCode",    "value": "USD"},
            {"type": "BOOLEAN",  "key": "remarketingOnly", "value": "false"},
            {"type": "BOOLEAN",  "key": "enabledMd",       "value": "false"}
        ],
        "firingTriggerId": cl_trigger_ids,
        "tagFiringOption": "ONCE_PER_EVENT",
        "monitoringMetadata": {"type": "MAP"},
        "consentSettings": {"consentStatus": "NOT_SET"}
    }


# ── Shared helpers (from generate_gtm_exports.py) ─────────────────────────────

def get_store_name(client_name, campaign_name=None):
    GENERIC    = {'main', 'primary', 'default', 'location', 'store', 'dfw', 'n/a'}
    SKIP_WORDS = {'auto', 'automotive', 'repair', 'service', 'center', 'care', 'shop',
                  'tire', 'garage', 'motors', 'motor', 'llc', 'inc', 'and', '&', 'the',
                  '1', 'of', 'at', 'in', 'for', 'n', 'by', 'st', 'ave', 'blvd'}
    name = str(client_name or '').strip()

    def _clean(s):
        s = re.sub(r'<[^>]+>', '', s)
        return s.replace('/', '').replace('  ', ' ').strip()

    def _valid(s):
        s = s.strip()
        if not s or s.lower() in GENERIC:
            return False
        if re.search(r'<[^>]+>', s):
            return False
        sl = s.lower()
        if sl.startswith('lnm') or sl in {'leads near me', 'general auto repair',
                                           'general auto repair services',
                                           'main campaign', 'general & euro'}:
            return False
        return True

    def _filter_words(text):
        return [w for w in text.split() if w.lower() not in SKIP_WORDS]

    if ' - ' in name:
        location = name.split(' - ')[-1].strip()
        cleaned  = _clean(location)
        if _valid(cleaned):
            return cleaned

    paren_match = re.search(r'\(([^)]+)\)', name)
    if paren_match:
        paren_words = _filter_words(paren_match.group(1))
        if paren_words and paren_words[0].lower() not in GENERIC and len(paren_words[0]) > 1:
            candidate = _clean(' '.join(paren_words[:3]))
            if _valid(candidate):
                return candidate

    if campaign_name and ' - ' in str(campaign_name):
        camp_city = str(campaign_name).split(' - ')[-1].strip()
        cleaned   = _clean(camp_city)
        if cleaned and _valid(cleaned):
            return cleaned

    name_no_paren = re.sub(r'\([^)]*\)', '', name).strip()
    words = _filter_words(name_no_paren)
    if words:
        return _clean(' '.join(words[:2]))

    return _clean(name_no_paren) or _clean(name)


def get_scheduler_info(scheduler_type):
    s = str(scheduler_type or '').lower()
    if 'shop genie' in s or 'shopgenie' in s:
        return 'appointment_booked', 'Shop Genie'
    if 'autoops' in s or 'auto ops' in s:
        return 'ao-appointment-booked', 'AutoOps'
    return 'dc-service-booked', 'OktoRocket'


def normalize_phone(raw):
    return re.sub(r'\D', '', str(raw or ''))


# ── Main ──────────────────────────────────────────────────────────────────────

def run(tier=None, row_filter=None, dry_run=False, force_recreate=False,
        rebuild_index=False, token_file=None):
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws_sheet = wb[SHEET_NAME]

    print("Building container index...")
    token_path = os.path.join(SCRIPT_DIR, token_file) if token_file else None
    service = get_gtm_service(token_path)

    # Derive a separate cache file for non-default token files
    if token_file:
        base = os.path.splitext(token_file)[0]  # e.g. "token_alex"
        cache_path = os.path.join(SCRIPT_DIR, f'container_index_cache_{base}.json')
    else:
        cache_path = None  # uses INDEX_CACHE default

    url_index, gtm_index = None, None
    if not rebuild_index:
        url_index, gtm_index = load_cached_index(cache_path)

    if url_index is None:
        url_index, gtm_index = build_container_index(service, cache_path)
    print()

    pushed       = 0
    skipped_nm   = 0   # no container match
    skipped_data = 0   # missing required data
    skipped_tier = 0
    no_match_rows = []

    rows = list(ws_sheet.iter_rows(min_row=2, values_only=True))

    for i, row in enumerate(rows, start=2):
        row_tier    = row[COL_TIER]
        client_name = row[COL_NAME]
        row_url     = row[COL_URL]
        sched       = row[COL_SCHED]
        phone_raw   = row[COL_PHONE]
        ga4_id      = row[COL_GA4_ID]
        gads_id     = row[COL_GADS_ID]
        appt_label  = row[COL_LABEL_APPT]
        phone_label = row[COL_LABEL_PHONE]
        row_gtm_id  = row[COL_GTM_PUBLIC] if len(row) > COL_GTM_PUBLIC else None
        multi_raw   = row[COL_MULTI_PHONE] if len(row) > COL_MULTI_PHONE else None

        if not client_name:
            continue

        # Row filter
        if row_filter is not None and i != row_filter:
            continue

        # Tier filter
        if tier is not None:
            try:
                if float(str(row_tier or '')) != float(tier):
                    skipped_tier += 1
                    continue
            except (ValueError, TypeError):
                skipped_tier += 1
                continue

        # Skip rows already marked done
        gtm_done_val = row[COL_GTM_DONE] if len(row) > COL_GTM_DONE else None
        if gtm_done_val and str(gtm_done_val).strip() not in ('', 'None', 'N/A', 'main site skip'):
            continue

        # Must have GA4 ID and GAds conversion data
        if not ga4_id or not gads_id or not appt_label:
            skipped_data += 1
            continue

        # Find container
        container = find_container(url_index, gtm_index, row_url, row_gtm_id)
        if container is None:
            domain = _normalize_domain(row_url)
            print(f"Row {i:3d}: '{client_name}' → NO CONTAINER MATCH (url={domain!r}, gtm={row_gtm_id!r})")
            no_match_rows.append((i, client_name, domain))
            skipped_nm += 1
            continue

        acct_id, ctr_id, public_id = container

        # Collect phone numbers
        phones = []
        p = normalize_phone(phone_raw)
        if p:
            phones.append(p)
        if multi_raw:
            try:
                for extra in json.loads(str(multi_raw)):
                    ep = normalize_phone(extra)
                    if ep and ep not in phones:
                        phones.append(ep)
            except Exception:
                pass

        has_phone  = bool(phones and phone_label)
        store_name = get_store_name(client_name)
        appt_event, sched_label = get_scheduler_info(sched)
        gads_id_str = str(int(float(str(gads_id))))

        domain = _normalize_domain(row_url)
        print(f"Row {i:3d}: '{client_name}' → {domain} → {public_id} "
              f"(acct={acct_id}/ctr={ctr_id})")

        if dry_run:
            # Show planned actions without making API calls
            print(f"  [dry-run] Trigger: CE - {sched_label} - Appointment Booked")
            for ph in phones:
                print(f"  [dry-run] Trigger: CL - Phone Click - {ph}")
            print(f"  [dry-run] Trigger: All Pages")
            print(f"  [dry-run] Tag: Conversion Linker")
            print(f"  [dry-run] Tag: GA4 - Configuration")
            print(f"  [dry-run] Tag: GA4 - Event - {appt_event}")
            if has_phone:
                print(f"  [dry-run] Tag: GA4 - Event - phone_click")
            print(f"  [dry-run] Tag: GAds - {store_name} - Booked_Appointment")
            if has_phone:
                print(f"  [dry-run] Tag: GAds - {store_name} - Phone_Click")
            pushed += 1
            continue

        # ── Live push ──────────────────────────────────────────────────────────
        try:
            workspace_id = get_workspace(service, acct_id, ctr_id)
            existing_triggers = get_existing_triggers(service, acct_id, ctr_id, workspace_id)
            existing_tags     = get_existing_tags(service, acct_id, ctr_id, workspace_id)

            # 1. Appointment trigger
            appt_body = build_appt_trigger_body(sched_label, appt_event)
            appt_trigger_id, appt_status = ensure_trigger(
                service, acct_id, ctr_id, workspace_id, appt_body,
                existing_triggers, force_recreate)
            print(f"  {'✓' if appt_status != 'existed' else '·'} Trigger: {appt_body['name']} ({appt_status})")

            # 2. Phone click triggers
            cl_trigger_ids = []
            for ph in phones:
                cl_body = build_cl_trigger_body(ph)
                cl_tid, cl_status = ensure_trigger(
                    service, acct_id, ctr_id, workspace_id, cl_body,
                    existing_triggers, force_recreate)
                cl_trigger_ids.append(cl_tid)
                print(f"  {'✓' if cl_status != 'existed' else '·'} Trigger: {cl_body['name']} ({cl_status})")

            # 3. All Pages trigger
            ap_body = build_all_pages_trigger_body()
            ap_trigger_id, ap_status = ensure_trigger(
                service, acct_id, ctr_id, workspace_id, ap_body,
                existing_triggers, force_recreate)
            print(f"  {'✓' if ap_status != 'existed' else '·'} Trigger: {ap_body['name']} ({ap_status})")

            # 4. Tags
            # Conversion Linker — must exist before any GAds tag fires
            tag_body = build_conversion_linker_tag_body(ap_trigger_id)
            _, ts = ensure_tag(service, acct_id, ctr_id, workspace_id,
                               tag_body, existing_tags, force_recreate)
            print(f"  {'✓' if ts != 'existed' else '·'} Tag: {tag_body['name']} ({ts})")

            # GA4 Config
            tag_body = build_ga4_config_tag_body(str(ga4_id), ap_trigger_id)
            _, ts = ensure_tag(service, acct_id, ctr_id, workspace_id,
                               tag_body, existing_tags, force_recreate)
            print(f"  {'✓' if ts != 'existed' else '·'} Tag: {tag_body['name']} ({ts})")

            # GA4 Event - appt
            tag_body = build_ga4_event_appt_tag_body(str(ga4_id), appt_event, appt_trigger_id)
            _, ts = ensure_tag(service, acct_id, ctr_id, workspace_id,
                               tag_body, existing_tags, force_recreate)
            print(f"  {'✓' if ts != 'existed' else '·'} Tag: {tag_body['name']} ({ts})")

            # GA4 Event - phone_click
            if has_phone:
                tag_body = build_ga4_event_phone_tag_body(str(ga4_id), cl_trigger_ids)
                _, ts = ensure_tag(service, acct_id, ctr_id, workspace_id,
                                   tag_body, existing_tags, force_recreate)
                print(f"  {'✓' if ts != 'existed' else '·'} Tag: {tag_body['name']} ({ts})")

            # GAds - Booked_Appointment
            tag_body = build_gads_appt_tag_body(store_name, gads_id_str, str(appt_label), appt_trigger_id)
            _, ts = ensure_tag(service, acct_id, ctr_id, workspace_id,
                               tag_body, existing_tags, force_recreate)
            print(f"  {'✓' if ts != 'existed' else '·'} Tag: {tag_body['name']} ({ts})")

            # GAds - Phone_Click
            if has_phone:
                tag_body = build_gads_phone_tag_body(store_name, gads_id_str, str(phone_label), cl_trigger_ids)
                _, ts = ensure_tag(service, acct_id, ctr_id, workspace_id,
                                   tag_body, existing_tags, force_recreate)
                print(f"  {'✓' if ts != 'existed' else '·'} Tag: {tag_body['name']} ({ts})")

            # Write-back GTM public ID to XLSX col 28 if missing
            if not row_gtm_id or str(row_gtm_id).strip() == '':
                _writeback_gtm_id(i, public_id)

            # Mark GTM setup complete with today's date
            if not dry_run:
                _writeback_gtm_done(i)

            pushed += 1

        except Exception as e:
            print(f"  ✗ ERROR: {e}")
            import traceback; traceback.print_exc()

    print(f"\n=== Done: {pushed} {'previewed' if dry_run else 'pushed'}, "
          f"{skipped_nm} skipped (no container match), "
          f"{skipped_data} skipped (missing data), "
          f"{skipped_tier} skipped (wrong tier) ===")

    if no_match_rows:
        print("\nNo-match rows:")
        for r, name, dom in no_match_rows:
            print(f"  Row {r}: '{name}' (domain={dom!r})")


def _writeback_gtm_id(row_num, public_id):
    """Write public_id into col 28 (1-based col 29) of the XLSX row."""
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME]
        ws.cell(row=row_num, column=COL_GTM_PUBLIC + 1).value = public_id
        wb.save(XLSX_PATH)
    except Exception as e:
        print(f"  [warn] Could not write GTM ID back to XLSX: {e}")


def _writeback_gtm_done(row_num):
    """Write today's date (YYYYMMDD) into col 19 (1-based col 20) of the XLSX row."""
    from datetime import date
    today = date.today().strftime('%Y%m%d')
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME]
        ws.cell(row=row_num, column=COL_GTM_DONE + 1).value = today
        wb.save(XLSX_PATH)
    except Exception as e:
        print(f"  [warn] Could not write GTM done date back to XLSX: {e}")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Push GTM tags/triggers via GTM API v2')
    parser.add_argument('--tier',           default=None, help='Only this tier (e.g. 2, 3, 3.5)')
    parser.add_argument('--row',            type=int, default=None, help='Single XLSX row number')
    parser.add_argument('--dry-run',        action='store_true', help='Preview matches + planned actions, no API calls')
    parser.add_argument('--force-recreate',  action='store_true', help='Delete and recreate existing tags/triggers')
    parser.add_argument('--rebuild-index',   action='store_true', help='Ignore cache and re-scan all GTM accounts')
    parser.add_argument('--token-file',      default=None, help='Token file to use (default: token.json). E.g. token_alex.json')
    args = parser.parse_args()

    run(
        tier=args.tier,
        row_filter=args.row,
        dry_run=args.dry_run,
        force_recreate=args.force_recreate,
        rebuild_index=args.rebuild_index,
        token_file=args.token_file,
    )
