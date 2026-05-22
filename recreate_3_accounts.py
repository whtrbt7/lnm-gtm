"""
Re-create GTM accounts for 3 specific rows that failed during the Playwright run.
Rows: 460, 525, 562
"""
import sys
sys.path.insert(0, '/Users/alexchiu/llmprojects/lnm-gtm')
from create_accounts import create_gtm_account
from playwright.sync_api import sync_playwright
import openpyxl, re, time, os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH  = os.path.join(SCRIPT_DIR, 'GTM Bulk Setup OktoRocket.xlsx')
SHEET_NAME = 'AA Client Import List (1)'
CDP_URL    = 'http://localhost:9222'

TARGET_ROWS = [534, 541, 573]
COL_NAME, COL_FOLDER, COL_URL, COL_GTM_ID = 1, 2, 20, 28


def clean_url(url):
    if not url:
        return ''
    return re.sub(r'^https?://', '', str(url)).rstrip('/')


def load_targets():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    targets = []
    for rn in TARGET_ROWS:
        row = list(ws.iter_rows(min_row=rn, max_row=rn, values_only=True))[0]
        name     = row[COL_NAME]
        folder   = row[COL_FOLDER]
        url      = row[COL_URL]
        gtm_id   = row[COL_GTM_ID]
        targets.append((rn, name, folder, clean_url(url), gtm_id))
    return targets


def writeback(row_num, gtm_id):
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]
    ws.cell(row=row_num, column=COL_GTM_ID + 1).value = gtm_id
    wb.save(XLSX_PATH)
    print(f'  Saved {gtm_id} -> row {row_num}')


def main():
    targets = load_targets()
    print('Targets:')
    for rn, name, folder, url, gtm_id in targets:
        status = f'already has {gtm_id}' if gtm_id else 'needs creation'
        account_name   = f'{name} - {folder}' if folder else name
        container_name = url or name
        print(f'  Row {rn}: "{account_name}" / "{container_name}" — {status}')

    to_create = [(rn, name, folder, url) for rn, name, folder, url, gtm_id in targets
                 if not (isinstance(gtm_id, str) and gtm_id.startswith('GTM-'))]
    print(f'\n{len(to_create)} to create\n')

    if not to_create:
        print('Nothing to do.')
        return

    with sync_playwright() as p:
        for attempt in range(10):
            try:
                browser = p.chromium.connect_over_cdp(CDP_URL)
                context = browser.contexts[0]
                break
            except Exception as e:
                print(f'  CDP attempt {attempt+1}: {e}')
                time.sleep(5)
        else:
            raise RuntimeError('Could not connect to Chrome')

        print('Connected to Chrome. Starting...\n')

        for rn, name, folder, url in to_create:
            account_name   = f'{name} - {folder}' if folder else name
            container_name = url or name
            print(f'Row {rn}: "{account_name}" / "{container_name}"')
            try:
                gtm_id = create_gtm_account(context, account_name, container_name)
                if gtm_id:
                    print(f'  SUCCESS: {gtm_id}')
                    writeback(rn, gtm_id)
                else:
                    print(f'  WARNING: No GTM ID captured for row {rn}')
            except Exception as e:
                print(f'  ERROR: {e}')
            time.sleep(2)

        browser.close()

    print('\nDone.')


if __name__ == '__main__':
    main()
